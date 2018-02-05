import json
from collections import OrderedDict
import logging

import math

from rules import Rules
from openpyxl import load_workbook
from data import Data


# with open('data.json') as file_data:
# with open('single_tree.json') as file_data:
#     data = json.load(file_data, object_pairs_hook=OrderedDict)


class RuleBase(object):
    def __init__(self, object_list, parent):
        self.obj_list = object_list
        self.intermediate_ref_val = 0
        self.rule_row_list = list()
        self.parent = parent
        self.con_ref_values = [0 for _ in range(len(self.parent.ref_val))]
        self.combinations = [[]]

    '''
    Create initial rule base
    '''

    def create_rule_base(self):
        cons_ref_val_1 = 0
        cons_ref_val_n = 0

        for each in self.obj_list:
            # calculate the range of consequence values(First and last)
            cons_ref_val_1 += float(
                float(each.attribute_weight) * float(each.ref_val[0])
            )
            cons_ref_val_n += float(
                float(each.attribute_weight) * float(each.ref_val[len(each.ref_val) - 1])
            )

        # logging.warning("Logged values {}, {}".format(cons_ref_val_1, cons_ref_val_n))

        self.con_ref_values[0] = cons_ref_val_1
        self.con_ref_values[len(self.con_ref_values) - 1] = cons_ref_val_n

        # calculate intermediate values within the range.
        intermediate_cons_ref_val_num = len(self.parent.ref_val) - 1
        for i in range(1, intermediate_cons_ref_val_num):
            current_val = float(cons_ref_val_1 * i * 1.0) + float(cons_ref_val_n * (intermediate_cons_ref_val_num - i))
            current_val /= (i + (intermediate_cons_ref_val_num - i))
            self.con_ref_values[i] = current_val
            # logging.warning("New value: {}".format(current_val))


        # Calculate the number of possible combinations of reference values.
        array_for_ref_count = [[]]

        for each in self.obj_list:
            length = len(each.ref_val)
            temp = list()
            for i in range(length):
                temp.append(i)
            array_for_ref_count.append(temp)

        array = array_for_ref_count[1:]
        pools = [tuple(pool) for pool in array] * 1
        result = [[]]
        for pool in pools:
            result = [x+[y] for x in result for y in pool]

        for each in result:
            self.combinations.append(each)

        self.combinations = self.combinations[1:]

        # print "Combinations are: {}".format(combinations)

        print (": {}".format(self.con_ref_values))

        # Calculate y for each combination and distribute consequence values in the range
        for each in self.combinations:
            rules = Rules()
            row_val = [0 for _ in range(len(self.con_ref_values))]
            y = 0.0
            for i in range(len(self.obj_list)):
                y += float(float(self.obj_list[i].ref_val[each[i]]) * float(self.obj_list[i].attribute_weight))
            # print "Y value: {}".format(y)

            is_continue = False
            for idx, per in enumerate(self.con_ref_values):
                if per == y:
                    row_val[idx] = 1.0
                    rules.consequence_val = row_val
                    rules.combinations = each
                    self.rule_row_list.append(rules)
                    is_continue = True
            if is_continue:
                continue
            else:
                for idx in range(len(self.con_ref_values) - 1):
                    if (self.con_ref_values[idx] > y) and (y > self.con_ref_values[idx + 1]):
                        row_val[idx + 1]  = float(
                                (self.con_ref_values[idx] - y) /
                                (self.con_ref_values[idx] - self.con_ref_values[idx + 1])
                            )
                        row_val[idx] = float(1 - row_val[idx + 1])

                rules.consequence_val = row_val
                rules.combinations = each
                self.rule_row_list.append(rules)

        print ("Rule Base: ")
        for each in self.rule_row_list:
            print (each.__dict__)
        return self.rule_row_list

    def generate_extended_belief_rule_base(self):
        wb = load_workbook('DataCenter.xlsx')
        ws = wb.active
        n = ws.max_row
        col = []

        for i in range(len(self.obj_list)):
            col.append(self.obj_list[i].name)
        for row in range(1, n+1):
            rule = Rules()
            # antecedent_dist = list()
            for idx, column in enumerate(col):
                cell_name = "{}{}".format(column, row)
                input_value = ws[cell_name].value
                self.input_transformation2(idx, input_value)
                # antecedent_dist.append(self.obj_list[idx].transformed_val)
                rule.antecedents_belief_dist.append(self.obj_list[idx].transformed_val)
            # rule.antecedents_belief_dist.append(antecedent_dist)
            parent_cell_name = "{}{}".format(self.parent.name, row)
            consequent_input_value = ws[parent_cell_name].value
            self.input_transformation2(None, consequent_input_value)
            rule.consequence_belief_dist = self.parent.transformed_val
            self.rule_row_list.append(rule)
        return self.rule_row_list

    def user_input_transformation(self, attribute_index, attribute_input_value):
        if attribute_index is None:
            attribute = self.parent
        else:
            attribute = self.obj_list[attribute_index]
        attribute.transformed_val = [0 for _ in range(len(attribute.ref_val))]
        for i in range(len(attribute.ref_val)-1):
            if (float(attribute.ref_val[i]) > attribute_input_value) and (attribute_input_value > float(attribute.ref_val[i + 1])):
                val_1 = (
                        (float(attribute.ref_val[i]) - attribute_input_value) / (float(attribute.ref_val[i]) - float(attribute.ref_val[i + 1]))
                )
                attribute.transformed_val[i+1] = val_1
                val_2 = 1 - val_1
                attribute.transformed_val[i] = val_2

    def individual_matching_degree(self):
        for rule_index, rule in enumerate(self.rule_row_list):
            # attribute_individual_matching = list()
            # antecedent_belief_dist = current_rule.antecedents_belief_dist
            for attribute_index, belief_dist in enumerate(rule.antecedents_belief_dist):
                # antecedent_attribute = self.obj_list[attribute_index]
                temp = 0
                for index, belief in enumerate(belief_dist):
                    temp += pow((self.obj_list[attribute_index].transformed_val[index]-belief), 2)
                individual_matching = 1 - math.sqrt(temp / 2)
                rule.attributes_individual_matching.append(individual_matching)
                # attribute_individual_matching.append(individual_matching)
            # rule.attribute_individual_matching.append(attribute_individual_matching)

    '''
    Transform input value in the range of consequent values
                each.transformed_val[j + 1] = str(val_1)
    '''

    def input_transformation(self):
        for each in self.obj_list:
            print ("Input value for {} is {}".format(each.name, each.input_val))
            # print "Value before input transformation: {}".format(each.transformed_val)
            try:
                user_input = float(each.input_val)
                # user_input = float(each.crisp_val)
            except:
                user_input = 0

            if user_input > float(each.ref_val[0]):
                user_input = float(each.ref_val[0])
            elif user_input < float(each.ref_val[len(each.ref_val) - 1]):
                user_input = float(each.ref_val[len(each.ref_val) - 1])
            flag = False
            for i in range(len(each.ref_val)):
                if user_input == float(each.ref_val[i]):
                    each.transformed_val[i] = 1
                    flag = True
                    break
            if not flag:
                for j in range(len(each.ref_val) - 1):
                    if (float(each.ref_val[j]) > user_input) and (user_input > float(each.ref_val[j+1])):
                        val_1 = (
                            (float(each.ref_val[j]) - user_input) / (float(each.ref_val[j]) - float(each.ref_val[j+1]))
                        )
                        each.transformed_val[j + 1] = val_1
                        val_2 = 1 - val_1
                        each.transformed_val[j] = val_2
            print("Value after input transformation: {}".format(each.transformed_val))

    def activation_weight_calculation(self):
        max_attribute_weight = 0
        for each in range(len(self.obj_list)):
            if self.obj_list[each].attribute_wight > max_attribute_weight:
                max_attribute_weight = self.obj_list[each].attribute_weight

        sum_matching_degree = 0
        for rule_index, rule in enumerate(self.rule_row_list):
            rule_matching_degree = 1
            for attribute_index, attribute in enumerate(self.obj_list):
                rule_matching_degree *= pow(rule.attributes_individual_matching[attribute_index], attribute.attribute_weight / max_attribute_weight)
            rule.matching_degree = rule_matching_degree *rule.rule_weight
            # rule.matching_degree = rule_matching_degree
            sum_matching_degree += rule_matching_degree

        for rule_index, rule in enumerate(self.rule_row_list):
            rule.activation_weight = rule.matching_degree/sum_matching_degree

    '''
    Calculate activation weight
    '''

    def activation_weight(self):
        # matching_degree = list()

        for i, row in enumerate(self.combinations):
            current_rule = self.rule_row_list[i]
            degree = 1.0
            for idx, val in enumerate(row):
                degree *= float(
                    pow(float(self.obj_list[idx].transformed_val[val]), float(self.obj_list[idx].attribute_weight))

                )
            # matching_degree.insert(i, degree)
            current_rule.matching_degree = degree

        sum = 0.0
        for k in range(len(self.rule_row_list)):
            current_rule = self.rule_row_list[k]
            # current_rule.matching_degree = matching_degree[k]
            sum += float(current_rule.rule_weight) * float(current_rule.matching_degree)

        for p in range(len(self.rule_row_list)):
            current_rule = self.rule_row_list[p]
            activation_weight = float(
                (float(current_rule.rule_weight) * float(current_rule.matching_degree)) /
                sum
            )
            current_rule.activation_weight = activation_weight

    '''
    Update rule base
    '''

    def belief_update(self):
        tao = [0 for _ in range(len(self.obj_list))]
        for i in range(len(self.obj_list)):
            # if obj_list[i].name != 'x8':
            try:
                input_val = float(self.obj_list[i].input_val)
            except:
                input_val = 0
            if input_val != 0:
                tao[i] = 1
        total = 0

        for j in range(len(self.obj_list)):
            summation = sum([float(each) for each in self.obj_list[j].transformed_val])
            total += summation

        if sum([each for each in tao]) <= 0:
            update_value = 1
        else:
            update_value = total / float(sum([each for each in tao]))

        for each in self.rule_row_list:
            new_val_list = []
            for idx, row in enumerate(each.consequence_val):
                # print "Before: {} {}".format(idx, row)
                new_val = float(row) * update_value
                new_val_list.insert(idx, new_val)
                # print "After: {} {}".format(idx, new_val)
            each.consequence_val = new_val_list

        for each in self.rule_row_list:
            print ("{}".format(each.consequence_val))

    '''
    Rule aggregation
    '''
    def rules_aggregation_anlytical_evidential_Reasoning(self):
        b = [0 for _ in range(len(self.rule_row_list))]
        a = [[0 for _ in range(len(self.con_ref_values))]for _ in range(len(self.rule_row_list))]
        c = [0 for _ in range(len(self.rule_row_list))]
        final_consequence = [0 for _ in range(len(self.con_ref_values))]
        product_a = 1
        product_b = [1 for _ in range(len(self.con_ref_values))]
        product_c = 1
        sum_product_b = 0

        for i in range(len(self.rule_row_list)):
            c[i] *= 1 - (float(self.rule_row_list[j].activation_weight))
            product_c *= c[i]
            for j in range(len(self.rule_row_list[i])):
                b[i] += self.rule_row_list[i][j]
                b[i] = 1-b[i]
            product_a *= b[i]
            for j in range(len(self.rule_row_list[i])):
                a[i][j] = self.rule_row_list[i][j]-b[i]+1
                product_b[j] *= a[i][j]
        for j in range(len(product_b)):
            sum_product_b += product_b[j]
        for j in range(len(product_b)):
            final_consequence[j] = (product_b[j]-product_a)/(sum_product_b-product_c-(len(self.con_ref_values))-1)

        return final_consequence

    def aggregate_rule(self):
        # Get all the consequent value list from rule base
        consequent_array = []

        for i in range(len(self.rule_row_list)):
            row = self.rule_row_list[i]
            consequent_array.insert(i, row.consequence_val)

        # Calculate mn from the consequent array and save in a 2D array(named mn here)
        mn = [[0 for _ in range(len(self.combinations))] for _ in range(len(self.con_ref_values))]

        for i in range(len(self.rule_row_list)):
            for idx, each in enumerate(self.rule_row_list[i].consequence_val):
                # import pdb;
                # pdb.set_trace()
                # for key, val in enumerate(each):
                mn[idx][i] = float(
                        float(self.rule_row_list[i].activation_weight) *
                        float(each)
                    )

        # Calculate md from the consequent array and save in a 1Ds array(named md here)
        md = [0 for _ in range(len(self.rule_row_list))]

        for j in range(len(consequent_array)):
            total = 0
            for k in range(len(consequent_array[j])):
                total += consequent_array[j][k]

            md[j] = 1 - (float(self.rule_row_list[j].activation_weight) * total)

        # Calculate d in a 1D array in several steps

        # Step 1: calculate total rowsum
        rowsum = [1 for _ in range(3)]

        for x in range(len(rowsum)):
            for y in range(len(self.rule_row_list)):
                rowsum[x] *= (mn[x][y] + md[y])

        total_rowsum = sum(rowsum)

        # Step 2: Calculate mh and save in a 1D array
        mh = 1

        for i in range(len(md)):
          mh *= md[i]

        # Step 3: Calculate kn, kn1, m(1D array), mhn and aggregated_consequence_val
        kn = total_rowsum - (2 * mh)
        kn1 = 1 / kn

        m = [0 for _ in range(3)]

        for j in range(3):
            m[j] = kn1 * (rowsum[j] - mh)

        mhn = kn1 * mh

        aggregated_consequence_val = [0 for k in range(3)]
        for k in range(len(rowsum)):
            aggregated_consequence_val[k] = m[k] / (1 - mhn)

        # output = [round(each, 2) for each in aggregated_consequence_val]
        output = [each for each in aggregated_consequence_val]
        print ("Aggregated Rules: ".format(output))
        return output


# obj_list = list()
# parent = ""
#
# for each in data:
#     obj = Data(**data[each])
#     obj.name = str(each)
#     # import pdb; pdb.set_trace()
#     if obj.parent != 'x12':
#     # if obj.parent != 'x8':
#         parent = obj
#     else:
#         obj_list.append(obj)
#
#
# print "Initial Data:"
# print "Antecedent ID    Name     Attribute Weight    Reference Titles    Reference Values"
# for row in obj_list:
#     print "{}   {}  {}  {}  {}".format(row.antecedent_id, row.antecedent_name, row.attribute_weight, row.ref_title, row.ref_val)
#
# rule_base = RuleBase(obj_list, parent)
# ref_val_list = rule_base.create_rule_base()
#
# # print "\n\n"
# # print "Rule Base: "
# # for each in ref_val_list:
# #     print "{} {} {} {} {} {}".format(each.antecedent_1, each.antecedent_1_ref_title, each.antecedent_2, each.antecedent_2_ref_title, each.rule_weight, each.consequence_val)
#
# print "\n\n"
# print "Input Transformation: "
#
# rule_base.input_transformation()
#
# print "\n\n"
# print "Rule base with activation weight"
#
# rule_base.activation_weight()
#
# print "\n\n"
# print "Belief Update"
# rule_base.belief_update()
#
# print "\n\n"
# print "Aggregated rule:"
# rule_base.aggregate_rule()
